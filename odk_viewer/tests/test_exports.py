from sys import stdout
import os
import datetime
import json
import StringIO
import csv
import tempfile
import zipfile
import shutil
from openpyxl import load_workbook
from time import sleep
from pyxform.builder import create_survey_from_xls
from django.conf import settings
from main.tests.test_base import MainTestCase
from django.utils.dateparse import parse_datetime
from django.core.urlresolvers import reverse
from django.core.files.temp import NamedTemporaryFile
from odk_viewer.xls_writer import XlsWriter
from odk_viewer.views import delete_export, export_list, create_export,\
    export_progress, export_download
from pyxform import SurveyElementBuilder
from odk_viewer.models import Export, ParsedInstance
from utils.export_tools import generate_export, increment_index_in_filename,\
    dict_to_joined_export, ExportBuilder
from odk_logger.models import Instance, XForm
from main.views import delete_data
from utils.logger_tools import inject_instanceid
from django.core.files.storage import get_storage_class
from odk_viewer.pandas_mongo_bridge import NoRecordsFoundError
from odk_viewer.tasks import create_xls_export
from xlrd import open_workbook
from odk_viewer.models.parsed_instance import _encode_for_mongo
from odk_logger.xform_instance_parser import XFormInstanceParser


class TestExportList(MainTestCase):
    def setUp(self):
        super(TestExportList, self).setUp()
        self._publish_transportation_form()
        survey = self.surveys[0]
        self._make_submission(
            os.path.join(
                self.this_directory, 'fixtures', 'transportation',
                'instances', survey, survey + '.xml'))

    def test_csv_export_list(self):
        kwargs = {'username': self.user.username,
                  'id_string': self.xform.id_string,
                  'export_type': Export.CSV_EXPORT}

        # test csv
        url = reverse(export_list, kwargs=kwargs)
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

    def test_xls_export_list(self):
        kwargs = {'username': self.user.username,
                  'id_string': self.xform.id_string,
                  'export_type': Export.XLS_EXPORT}
        url = reverse(export_list, kwargs=kwargs)
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

    def test_kml_export_list(self):
        kwargs = {'username': self.user.username,
                  'id_string': self.xform.id_string,
                  'export_type': Export.KML_EXPORT}
        url = reverse(export_list, kwargs=kwargs)
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

    def test_zip_export_list(self):
        kwargs = {'username': self.user.username,
                  'id_string': self.xform.id_string,
                  'export_type': Export.ZIP_EXPORT}
        url = reverse(export_list, kwargs=kwargs)
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

    def test_gdoc_export_list(self):
        kwargs = {'username': self.user.username,
                  'id_string': self.xform.id_string,
                  'export_type': Export.GDOC_EXPORT}
        url = reverse(export_list, kwargs=kwargs)
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)

    def test_xsv_zip_export_list(self):
        kwargs = {'username': self.user.username,
                  'id_string': self.xform.id_string,
                  'export_type': Export.CSV_ZIP_EXPORT}
        url = reverse(export_list, kwargs=kwargs)
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)


class TestDataExportURL(MainTestCase):
    def setUp(self):
        super(TestDataExportURL, self).setUp()
        self._publish_transportation_form()

    def _filename_from_disposition(self, content_disposition):
        filename_pos = content_disposition.index('filename=')
        self.assertTrue(filename_pos != -1)
        return content_disposition[filename_pos + len('filename='):]

    def test_csv_export_url(self):
        self._submit_transport_instance()
        url = reverse('csv_export', kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
        })
        response = self.client.get(url)
        headers = dict(response.items())
        self.assertEqual(headers['Content-Type'], 'application/csv')
        content_disposition = headers['Content-Disposition']
        filename = self._filename_from_disposition(content_disposition)
        basename, ext = os.path.splitext(filename)
        self.assertEqual(ext, '.csv')

    def test_csv_export_url_without_records(self):
        # csv using the pandas path can throw a NoRecordsFound Exception -
        # handle it gracefully
        url = reverse('csv_export', kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
        })
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)

    def test_xls_export_url(self):
        self._submit_transport_instance()
        url = reverse('xls_export', kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
        })
        response = self.client.get(url)
        headers = dict(response.items())
        self.assertEqual(headers['Content-Type'],
                         'application/vnd.openxmlformats')
        content_disposition = headers['Content-Disposition']
        filename = self._filename_from_disposition(content_disposition)
        basename, ext = os.path.splitext(filename)
        self.assertEqual(ext, '.xlsx')

    def test_csv_zip_export_url(self):
        self._submit_transport_instance()
        url = reverse('csv_zip_export', kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
        })
        response = self.client.get(url)
        headers = dict(response.items())
        self.assertEqual(headers['Content-Type'], 'application/zip')
        content_disposition = headers['Content-Disposition']
        filename = self._filename_from_disposition(content_disposition)
        basename, ext = os.path.splitext(filename)
        self.assertEqual(ext, '.zip')


class TestExports(MainTestCase):
    def setUp(self):
        super(TestExports, self).setUp()
        self._submission_time = parse_datetime('2013-02-18 15:54:01Z')

    def test_unique_xls_sheet_name(self):
        xls_writer = XlsWriter()
        xls_writer.add_sheet('section9_pit_latrine_with_slab_group')
        xls_writer.add_sheet('section9_pit_latrine_without_slab_group')
        # create a set of sheet names keys
        sheet_names_set = set(xls_writer._sheets.keys())
        self.assertEqual(len(sheet_names_set), 2)

    def test_csv_http_response(self):
        self._publish_transportation_form()
        survey = self.surveys[0]
        self._make_submission(
            os.path.join(
                self.this_directory, 'fixtures', 'transportation',
                'instances', survey, survey + '.xml'),
            forced_submission_time=self._submission_time)
        response = self.client.get(reverse('csv_export',
            kwargs={
                'username': self.user.username,
                'id_string': self.xform.id_string
            }))
        self.assertEqual(response.status_code, 200)
        test_file_path = os.path.join(os.path.dirname(__file__),
            'fixtures', 'transportation.csv')
        content = self._get_response_content(response)
        with open(test_file_path, 'r') as test_file:
            self.assertEqual(content, test_file.read())

    def test_responses_for_empty_exports(self):
        self._publish_transportation_form()
        # test csv though xls uses the same view
        url = reverse('csv_export',
            kwargs={
                'username': self.user.username,
                'id_string': self.xform.id_string
            }
        )
        self.response = self.client.get(url)
        self.assertEqual(self.response.status_code, 404)
        self.assertIn('text/html', self.response['content-type'])

    def test_create_export(self):
        self._publish_transportation_form_and_submit_instance()
        storage = get_storage_class()()
        # test xls
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
            self.xform.id_string)
        self.assertTrue(storage.exists(export.filepath))
        path, ext = os.path.splitext(export.filename)
        self.assertEqual(ext, '.xls')

        # test csv
        export = generate_export(Export.CSV_EXPORT, 'csv', self.user.username,
            self.xform.id_string)
        self.assertTrue(storage.exists(export.filepath))
        path, ext = os.path.splitext(export.filename)
        self.assertEqual(ext, '.csv')

        # test xls with existing export_id
        existing_export = Export.objects.create(xform=self.xform,
            export_type=Export.XLS_EXPORT)
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
            self.xform.id_string, existing_export.id)
        self.assertEqual(existing_export.id, export.id)

    def test_delete_file_on_export_delete(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
                                 self.xform.id_string)
        storage = get_storage_class()()
        self.assertTrue(storage.exists(export.filepath))
        # delete export object
        export.delete()
        self.assertFalse(storage.exists(export.filepath))

    def test_graceful_exit_on_export_delete_if_file_doesnt_exist(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
            self.xform.id_string)
        storage = get_storage_class()()
        # delete file
        storage.delete(export.filepath)
        self.assertFalse(storage.exists(export.filepath))
        # clear filename, like it would be in an incomplete export
        export.filename = None
        export.filedir = None
        export.save()
        # delete export record, which should try to delete file as well
        delete_url = reverse(delete_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        post_data = {'export_id': export.id}
        response = self.client.post(delete_url, post_data)
        self.assertEqual(response.status_code, 302)

    def test_delete_oldest_export_on_limit(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create first export
        first_export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
            self.xform.id_string)
        self.assertIsNotNone(first_export.pk)
        # create exports that exceed set limit
        for i in range(Export.MAX_EXPORTS):
            generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
                self.xform.id_string)
        # first export should be deleted
        exports = Export.objects.filter(id=first_export.id)
        self.assertEqual(len(exports), 0)

    def test_create_export_url(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        num_exports = Export.objects.count()
        # create export
        create_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.XLS_EXPORT
        })
        response = self.client.post(create_export_url)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Export.objects.count(), num_exports + 1)

    def test_delete_export_url(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create export
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
            self.xform.id_string)
        exports = Export.objects.filter(id=export.id)
        self.assertEqual(len(exports), 1)
        delete_url = reverse(delete_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        post_data = {'export_id': export.id}
        response = self.client.post(delete_url, post_data)
        self.assertEqual(response.status_code, 302)
        exports = Export.objects.filter(id=export.id)
        self.assertEqual(len(exports), 0)

    def test_export_progress_output(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create exports
        for i in range(2):
            generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
                self.xform.id_string)
        self.assertEqual(Export.objects.count(), 2)
        # progress for multiple exports
        progress_url = reverse(export_progress, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        get_data = {'export_ids': [e.id for e in Export.objects.all()]}
        response = self.client.get(progress_url, get_data)
        content = json.loads(response.content)
        self.assertEqual(len(content), 2)
        self.assertEqual(sorted(['url', 'export_id', 'complete', 'filename']),
            sorted(content[0].keys()))

    def test_auto_export_if_none_exists(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # get export list url
        num_exports = Export.objects.count()
        export_list_url = reverse(export_list, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.XLS_EXPORT
        })
        response = self.client.get(export_list_url)
        self.assertEqual(Export.objects.count(), num_exports + 1)

    def test_dont_auto_export_if_exports_exist(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create export
        create_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.XLS_EXPORT
        })
        response = self.client.post(create_export_url)
        num_exports = Export.objects.count()
        export_list_url = reverse(export_list, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.XLS_EXPORT
        })
        response = self.client.get(export_list_url)
        self.assertEqual(Export.objects.count(), num_exports)

    def test_last_submission_time_on_export(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create export
        xls_export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
            self.xform.id_string)
        num_exports = Export.objects.filter(xform=self.xform,
            export_type=Export.XLS_EXPORT).count()
        # check that our function knows there are no more submissions
        self.assertFalse(Export.exports_outdated(xform=self.xform,
            export_type=Export.XLS_EXPORT))
        sleep(1)
        # force new  last submission date on xform
        last_submission = self.xform.surveys.order_by('-date_created')[0]
        last_submission.date_created += datetime.timedelta(hours=1)
        last_submission.save()
        # check that our function knows data has changed
        self.assertTrue(Export.exports_outdated(xform=self.xform,
            export_type=Export.XLS_EXPORT))
        # check that requesting list url will generate a new export
        export_list_url = reverse(export_list, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.XLS_EXPORT
        })
        response = self.client.get(export_list_url)
        self.assertEqual(Export.objects.filter(xform=self.xform,
            export_type=Export.XLS_EXPORT).count(), num_exports + 1)
        # make sure another export type causes auto-generation
        num_exports = Export.objects.filter(xform=self.xform,
            export_type=Export.CSV_EXPORT).count()
        export_list_url = reverse(export_list, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': Export.CSV_EXPORT
        })
        response = self.client.get(export_list_url)
        self.assertEqual(Export.objects.filter(xform=self.xform,
            export_type=Export.CSV_EXPORT).count(), num_exports + 1)

    def test_last_submission_time_empty(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create export
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
            self.xform.id_string)
        # set time of last submission to None
        export.time_of_last_submission = None
        export.save()
        self.assertTrue(Export.exports_outdated(xform=self.xform,
            export_type=Export.XLS_EXPORT))

    def test_invalid_export_type(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        export_list_url = reverse(export_list, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'invalid'
        })
        response = self.client.get(export_list_url)
        self.assertEqual(response.status_code, 400)
        # test create url
        create_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'invalid'
        })
        response = self.client.post(create_export_url)
        self.assertEqual(response.status_code, 400)

    def test_add_index_to_filename(self):
        filename = "file_name-123f.txt"
        new_filename = increment_index_in_filename(filename)
        expected_filename = "file_name-123f-1.txt"
        self.assertEqual(new_filename, expected_filename)

        # test file that already has an index
        filename = "file_name-123.txt"
        new_filename = increment_index_in_filename(filename)
        expected_filename = "file_name-124.txt"
        self.assertEqual(new_filename, expected_filename)

    def test_duplicate_export_filename_is_renamed(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create an export object in the db
        # TODO: only works if the time we time we generate the basename is exact to the second with the time the 2nd export is created
        basename = "%s_%s" % (self.xform.id_string,
                              datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S"))
        filename = basename + ".csv"
        export = Export.objects.create(xform=self.xform,
            export_type=Export.CSV_EXPORT, filename=filename)
        # 2nd export
        export_2 = generate_export(Export.CSV_EXPORT, 'csv', self.user.username,
                                   self.xform.id_string)
        if export.created_on.timetuple() == export_2.created_on.timetuple():
            new_filename = increment_index_in_filename(filename)
            self.assertEqual(new_filename, export_2.filename)
        else:
            stdout.write("duplicate export filename test skipped because export times differ.")

    def test_export_download_url(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        export = generate_export(Export.CSV_EXPORT, 'csv', self.user.username,
                                 self.xform.id_string)
        csv_export_url = reverse(export_download, kwargs={
            "username": self.user.username,
            "id_string": self.xform.id_string,
            "export_type": Export.CSV_EXPORT,
            "filename": export.filename
        })
        response = self.client.get(csv_export_url)
        self.assertEqual(response.status_code, 200)
        # test xls
        export = generate_export(Export.XLS_EXPORT, 'xls', self.user.username,
            self.xform.id_string)
        xls_export_url = reverse(export_download, kwargs={
            "username": self.user.username,
            "id_string": self.xform.id_string,
            "export_type": Export.XLS_EXPORT,
            "filename": export.filename
        })
        response = self.client.get(xls_export_url)
        self.assertEqual(response.status_code, 200)

    def test_404_on_export_io_error(self):
        """
        Test that we return a 404 when the response_with_mimetype_and_name encounters an IOError
        """
        self._publish_transportation_form()
        self._submit_transport_instance()
        export = generate_export(Export.CSV_EXPORT, 'csv', self.user.username,
                                 self.xform.id_string)
        export_url = reverse(export_download, kwargs={
            "username": self.user.username,
            "id_string": self.xform.id_string,
            "export_type": Export.CSV_EXPORT,
            "filename": export.filename
        })
        # delete the export
        export.delete()
        # access the export
        response = self.client.get(export_url)
        self.assertEqual(response.status_code, 404)

    def test_deleted_submission_not_in_export(self):
        self._publish_transportation_form()
        initial_count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        self._submit_transport_instance(0)
        self._submit_transport_instance(1)
        count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        self.assertEqual(count, initial_count+2)
        # get id of second submission
        instance_id = Instance.objects.filter(
            xform=self.xform).order_by('id').reverse()[0].id
        delete_url = reverse(
            delete_data, kwargs={"username": self.user.username,
                                 "id_string": self.xform.id_string})
        params = {'id': instance_id}
        self.client.post(delete_url, params)
        count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        self.assertEqual(count, initial_count + 1)
        # create the export
        csv_export_url = reverse(
            'csv_export', kwargs={"username": self.user.username,
                                "id_string":self.xform.id_string})
        response = self.client.get(csv_export_url)
        self.assertEqual(response.status_code, 200)
        f = StringIO.StringIO(self._get_response_content(response))
        csv_reader = csv.reader(f)
        num_rows = len([row for row in csv_reader])
        f.close()
        # number of rows == 2 i.e. initial_count + header plus one row
        self.assertEqual(num_rows, initial_count + 2)

    def test_edited_submissions_in_exports(self):
        self._publish_transportation_form()
        initial_count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        instance_name = 'transport_2011-07-25_19-05-36'
        path = os.path.join(
            'main', 'tests', 'fixtures', 'transportation', 'instances_w_uuid',
            instance_name, instance_name + '.xml')
        self._make_submission(path)
        count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        self.assertEqual(count, initial_count+1)
        instance = Instance.objects.filter(
            xform=self.xform).order_by('id').reverse()[0]
        # make edited submission - simulating what enketo would return
        instance_name = 'transport_2011-07-25_19-05-36-edited'
        path = os.path.join(
            'main', 'tests', 'fixtures', 'transportation', 'instances_w_uuid',
            instance_name, instance_name + '.xml')
        self._make_submission(path)
        count = ParsedInstance.query_mongo(
            self.user.username, self.xform.id_string, '{}', '[]', '{}',
            count=True)[0]['count']
        self.assertEqual(count, initial_count+1)
        # create the export
        csv_export_url = reverse(
            'csv_export', kwargs={"username": self.user.username,
                                "id_string":self.xform.id_string})
        response = self.client.get(csv_export_url)
        self.assertEqual(response.status_code, 200)
        f = StringIO.StringIO(self._get_response_content(response))
        csv_reader = csv.DictReader(f)
        data = [row for row in csv_reader]
        f.close()
        num_rows = len(data)
        # number of rows == initial_count + 1
        self.assertEqual(num_rows, initial_count + 1)
        key ='transport/loop_over_transport_types_frequency/ambulance/frequency_to_referral_facility'
        self.assertEqual(data[initial_count][key], "monthly")

    def test_export_ids_dont_have_comma_separation(self):
        """
        It seems using {{ }} to output numbers greater than 1000 formats the
        number with a thousand separator
        """
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create an in-complete export
        export = Export.objects.create(id=1234, xform=self.xform,
                              export_type=Export.XLS_EXPORT)
        self.assertEqual(export.pk, 1234)
        export_list_url = reverse(
            export_list, kwargs={
                "username": self.user.username,
                "id_string": self.xform.id_string,
                "export_type": Export.XLS_EXPORT
            })
        response = self.client.get(export_list_url)
        self.assertContains(response, '#delete-1234"')
        self.assertNotContains(response, '#delete-1,234"')

    def test_export_progress_updates(self):
        """
        Test that after generate_export is called, we change out state to
        pending and after its complete, we change it to complete, if we fail
        between the two, updates, we have failed
        """
        self._publish_transportation_form()
        # generate an export that fails because of the NoRecordsFound exception
        export = Export.objects.create(xform=self.xform,
            export_type=Export.XLS_EXPORT)
        # check that progress url says pending
        progress_url = reverse(export_progress, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        params = {'export_ids': [export.id]}
        response = self.client.get(progress_url, params)
        status = json.loads(response.content)[0]
        self.assertEqual(status["complete"], False)
        self.assertEqual(status["filename"], None)

        export.internal_status = Export.FAILED
        export.save()
        # check that progress url says failed
        progress_url = reverse(export_progress, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        params = {'export_ids': [export.id]}
        response = self.client.get(progress_url, params)
        status = json.loads(response.content)[0]
        self.assertEqual(status["complete"], True)
        self.assertEqual(status["filename"], None)

        # make a submission and create a valid export
        self._submit_transport_instance()
        create_xls_export(
            self.user.username,
            self.xform.id_string, export.id)
        params = {'export_ids': [export.id]}
        response = self.client.get(progress_url, params)
        status = json.loads(response.content)[0]
        self.assertEqual(status["complete"], True)
        self.assertIsNotNone(status["filename"])

    def test_direct_export_returns_newset_export_if_not_updated_since(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        self.assertEqual(self.response.status_code, 201)
        sleep(1)
        self._submit_transport_instance_w_uuid("transport_2011-07-25_19-05-36")
        self.assertEqual(self.response.status_code, 201)

        initial_num_csv_exports = Export.objects.filter(
            xform=self.xform, export_type=Export.CSV_EXPORT).count()
        initial_num_xls_exports = Export.objects.filter(
            xform=self.xform, export_type=Export.XLS_EXPORT).count()
        # request a direct csv export
        csv_export_url = reverse('csv_export', kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string
        })
        xls_export_url = reverse('xls_export', kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string
        })
        response = self.client.get(csv_export_url)
        self.assertEqual(response.status_code, 200)
        # we should have initial_num_exports + 1 exports
        num_csv_exports = Export.objects.filter(
            xform=self.xform, export_type=Export.CSV_EXPORT).count()
        self.assertEqual(num_csv_exports, initial_num_csv_exports + 1)

        # request another export without changing the data
        response = self.client.get(csv_export_url)
        self.assertEqual(response.status_code, 200)
        # we should still only have a single export object
        num_csv_exports = Export.objects.filter(
            xform=self.xform, export_type=Export.CSV_EXPORT).count()
        self.assertEqual(num_csv_exports, initial_num_csv_exports + 1)

        # this should not affect a direct XLS export and XLS should still re-generate
        response = self.client.get(xls_export_url)
        self.assertEqual(response.status_code, 200)
        num_xls_exports = Export.objects.filter(
            xform=self.xform, export_type=Export.XLS_EXPORT).count()
        self.assertEqual(num_xls_exports, initial_num_xls_exports + 1)

        # make sure xls doesnt re-generate if data hasn't changed
        response = self.client.get(xls_export_url)
        self.assertEqual(response.status_code, 200)
        num_xls_exports = Export.objects.filter(
            xform=self.xform, export_type=Export.XLS_EXPORT).count()
        self.assertEqual(num_xls_exports, initial_num_xls_exports + 1)

        sleep(1)
        # check that data edits cause a re-generation
        self._submit_transport_instance_w_uuid(
            "transport_2011-07-25_19-05-36-edited")
        self.assertEqual(self.response.status_code, 201)
        self.client.get(csv_export_url)
        self.assertEqual(response.status_code, 200)
        # we should have an extra export now that the data has been updated
        num_csv_exports = Export.objects.filter(
            xform=self.xform, export_type=Export.CSV_EXPORT).count()
        self.assertEqual(num_csv_exports, initial_num_csv_exports + 2)

        sleep(1)
        # and when we delete
        delete_url = reverse(delete_data, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string
        })
        instance = Instance.objects.filter().order_by('-pk')[0]
        response = self.client.post(delete_url, {'id': instance.id})
        self.assertEqual(response.status_code, 200)
        response = self.client.get(csv_export_url)
        self.assertEqual(response.status_code, 200)
        # we should have an extra export now that the data has been updated by the delete
        num_csv_exports = Export.objects.filter(
            xform=self.xform, export_type=Export.CSV_EXPORT).count()
        self.assertEqual(num_csv_exports, initial_num_csv_exports + 3)

    def test_exports_outdated_doesnt_consider_failed_exports(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create a bad export
        export = Export.objects.create(
            xform=self.xform, export_type=Export.XLS_EXPORT,
            internal_status=Export.FAILED)
        self.assertTrue(
            Export.exports_outdated(self.xform, export.export_type))

    def test_exports_outdated_considers_pending_exports(self):
        self._publish_transportation_form()
        self._submit_transport_instance()
        # create a pending export
        export = Export.objects.create(
            xform=self.xform, export_type=Export.XLS_EXPORT,
            internal_status=Export.PENDING)
        self.assertFalse(
            Export.exports_outdated(self.xform, export.export_type))

    def _get_csv_data(self, filepath):
        storage = get_storage_class()()
        csv_file = storage.open(filepath)
        reader = csv.DictReader(csv_file)
        data = reader.next()
        csv_file.close()
        return data

    def _get_xls_data(self, filepath):
        storage = get_storage_class()()
        with storage.open(filepath) as f:
            workbook = open_workbook(file_contents=f.read())
        transportation_sheet = workbook.sheet_by_name("transportation")
        self.assertTrue(transportation_sheet.nrows > 1)
        headers = transportation_sheet.row_values(0)
        column1 = transportation_sheet.row_values(1)
        return dict(zip(headers, column1))

    def test_column_header_delimiter_export_option(self):
        self._publish_transportation_form()
        # survey 1 has ambulance and bicycle as values for
        # transport/available_transportation_types_to_referral_facility
        self._submit_transport_instance(survey_at=1)
        create_csv_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'csv'
        })
        default_params = {}
        custom_params = {
            'options[group_delimiter]': '.',
        }
        # test csv with default group delimiter
        response = self.client.post(create_csv_export_url, default_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='csv').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_csv_data(export.filepath)
        self.assertTrue(
            data.has_key(
                'transport/available_transportation_types_to_referral_facility/ambulance'))
        self.assertEqual(
            data['transport/available_transportation_types_to_referral_facility/ambulance'], 'True')

        sleep(1)
        # test csv with dot delimiter
        response = self.client.post(create_csv_export_url, custom_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='csv').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_csv_data(export.filepath)
        self.assertTrue(
            data.has_key(
                'transport.available_transportation_types_to_referral_facility.ambulance'))
        self.assertEqual(
            data['transport.available_transportation_types_to_referral_facility.ambulance'], 'True')

        # test xls with default group delimiter
        create_csv_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        response = self.client.post(create_csv_export_url, default_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='xls').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_xls_data(export.filepath)
        self.assertTrue(
            data.has_key("transport/available_transportation_types_to_referral_facility/ambulance"))
        # xlrd reader seems to convert bools into integers i.e. 0 or 1
        self.assertEqual(
            data["transport/available_transportation_types_to_referral_facility/ambulance"], 1)

        sleep(1)
        # test xls with dot delimiter
        response = self.client.post(create_csv_export_url, custom_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='xls').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_xls_data(export.filepath)
        self.assertTrue(
            data.has_key("transport.available_transportation_types_to_referral_facility.ambulance"))
        # xlrd reader seems to convert bools into integers i.e. 0 or 1
        self.assertEqual(
            data["transport.available_transportation_types_to_referral_facility.ambulance"], 1)

    def test_split_select_multiple_export_option(self):
        self._publish_transportation_form()
        self._submit_transport_instance(survey_at=1)
        create_csv_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'csv'
        })
        default_params = {}
        custom_params = {
            'options[dont_split_select_multiples]': 'yes'
        }
        # test csv with default split select multiples
        response = self.client.post(create_csv_export_url, default_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='csv').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_csv_data(export.filepath)
        # we should have transport/available_transportation_types_to_referral_facility/ambulance as a separate column
        self.assertTrue(
            data.has_key(
                'transport/available_transportation_types_to_referral_facility/ambulance'))

        sleep(1)
        # test csv without default split select multiples
        response = self.client.post(create_csv_export_url, custom_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='csv').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_csv_data(export.filepath)
        # transport/available_transportation_types_to_referral_facility/ambulance should not be in its own column
        self.assertFalse(
            data.has_key(
                'transport/available_transportation_types_to_referral_facility/ambulance'))
        # transport/available_transportation_types_to_referral_facility should be a column
        self.assertTrue(
            data.has_key(
                'transport/available_transportation_types_to_referral_facility'))
        # check that ambulance is one the values within the transport/available_transportation_types_to_referral_facility column
        self.assertTrue("ambulance" in data['transport/available_transportation_types_to_referral_facility'].split(" "))

        create_xls_export_url = reverse(create_export, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': 'xls'
        })
        # test xls with default split select multiples
        response = self.client.post(create_xls_export_url, default_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='xls').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_xls_data(export.filepath)
        # we should have transport/available_transportation_types_to_referral_facility/ambulance as a separate column
        self.assertTrue(
            data.has_key(
                'transport/available_transportation_types_to_referral_facility/ambulance'))

        sleep(1)
        # test xls without default split select multiples
        response = self.client.post(create_xls_export_url, custom_params)
        self.assertEqual(response.status_code, 302)
        export = Export.objects.filter(
            xform=self.xform, export_type='xls').latest('created_on')
        self.assertTrue(bool(export.filepath))
        data = self._get_xls_data(export.filepath)
        # transport/available_transportation_types_to_referral_facility/ambulance should NOT be in its own column
        self.assertFalse(
            data.has_key(
                'transport/available_transportation_types_to_referral_facility/ambulance'))
        # transport/available_transportation_types_to_referral_facility should be a column
        self.assertTrue(
            data.has_key(
                'transport/available_transportation_types_to_referral_facility'))
        # check that ambulance is one the values within the transport/available_transportation_types_to_referral_facility column
        self.assertTrue("ambulance" in data['transport/available_transportation_types_to_referral_facility'].split(" "))

    def test_dict_to_joined_export_works(self):
        data =\
            {
                'name': 'Abe',
                'age': '35',
                '_geolocation': [None, None],
                'attachments': ['abcd.jpg', 'efgh.jpg'],
                'children':
                [
                    {
                        'children/name': 'Mike',
                        'children/age': '5',
                        'children/cartoons':
                        [
                            {
                                'children/cartoons/name': 'Tom & Jerry',
                                'children/cartoons/why': 'Tom is silly',
                            },
                            {
                                'children/cartoons/name': 'Flinstones',
                                'children/cartoons/why': u"I like bamb bam\u0107",
                            }
                        ]
                    },
                    {
                        'children/name': 'John',
                        'children/age': '2',
                        'children/cartoons':[]
                    },
                    {
                        'children/name': 'Imora',
                        'children/age': '3',
                        'children/cartoons':
                        [
                            {
                                'children/cartoons/name': 'Shrek',
                                'children/cartoons/why': 'He\'s so funny'
                            },
                            {
                                'children/cartoons/name': 'Dexter\'s Lab',
                                'children/cartoons/why': 'He thinks hes smart',
                                'children/cartoons/characters':
                                [
                                    {
                                        'children/cartoons/characters/name': 'Dee Dee',
                                        'children/cartoons/characters/good_or_evil': 'good'
                                    },
                                    {
                                        'children/cartoons/characters/name': 'Dexter',
                                        'children/cartoons/characters/good_or_evil': 'evil'
                                    },
                                ]
                            }
                        ]
                    }
                ]
            }
        expected_output =\
            {
                'survey': {
                  'name': 'Abe',
                  'age': '35'
                },
                'children':
                [
                    {
                        'children/name': 'Mike',
                        'children/age': '5',
                        '_index': 1,
                        '_parent_table_name': 'survey',
                        '_parent_index': 1
                    },
                    {
                        'children/name': 'John',
                        'children/age': '2',
                        '_index': 2,
                        '_parent_table_name': 'survey',
                        '_parent_index': 1
                    },
                    {
                        'children/name': 'Imora',
                        'children/age': '3',
                        '_index': 3,
                        '_parent_table_name': 'survey',
                        '_parent_index': 1
                    },
                ],
                'children/cartoons':
                [
                    {
                        'children/cartoons/name': 'Tom & Jerry',
                        'children/cartoons/why': 'Tom is silly',
                        '_index': 1,
                        '_parent_table_name': 'children',
                        '_parent_index': 1
                    },
                    {
                        'children/cartoons/name': 'Flinstones',
                        'children/cartoons/why': u"I like bamb bam\u0107",
                        '_index': 2,
                        '_parent_table_name': 'children',
                        '_parent_index': 1
                    },
                    {
                        'children/cartoons/name': 'Shrek',
                        'children/cartoons/why': 'He\'s so funny',
                        '_index': 3,
                        '_parent_table_name': 'children',
                        '_parent_index': 3
                    },
                    {
                        'children/cartoons/name': 'Dexter\'s Lab',
                        'children/cartoons/why': 'He thinks hes smart',
                        '_index': 4,
                        '_parent_table_name': 'children',
                        '_parent_index': 3
                    }
                ],
                'children/cartoons/characters':
                [
                    {
                        'children/cartoons/characters/name': 'Dee Dee',
                        'children/cartoons/characters/good_or_evil': 'good',
                        '_index': 1,
                        '_parent_table_name': 'children/cartoons',
                        '_parent_index': 4
                    },
                    {
                        'children/cartoons/characters/name': 'Dexter',
                        'children/cartoons/characters/good_or_evil': 'evil',
                        '_index': 2,
                        '_parent_table_name': 'children/cartoons',
                        '_parent_index': 4
                    }
                ]
            }
        survey_name = 'survey'
        indices = {survey_name: 0}
        output = dict_to_joined_export(data, 1, indices, survey_name)
        self.assertEqual(output[survey_name], expected_output[survey_name])
        # 1st level
        self.assertEqual(len(output['children']), 3)
        for child in enumerate(['Mike', 'John', 'Imora']):
            index = child[0]
            name = child[1]
            self.assertEqual(
                filter(
                    lambda x: x['children/name'] == name,
                    output['children'])[0],
                expected_output['children'][index])
        # 2nd level
        self.assertEqual(len(output['children/cartoons']), 4)
        for cartoon in enumerate(
                ['Tom & Jerry', 'Flinstones', 'Shrek', 'Dexter\'s Lab']):
            index = cartoon[0]
            name = cartoon[1]
            self.assertEqual(
                filter(
                    lambda x: x['children/cartoons/name'] == name,
                    output['children/cartoons'])[0],
                expected_output['children/cartoons'][index])
        # 3rd level
        self.assertEqual(len(output['children/cartoons/characters']), 2)
        for characters in enumerate(['Dee Dee', 'Dexter']):
            index = characters[0]
            name = characters[1]
            self.assertEqual(
                filter(
                    lambda x: x['children/cartoons/characters/name'] == name,
                    output['children/cartoons/characters'])[0],
                expected_output['children/cartoons/characters'][index])

    def test_generate_csv_zip_export(self):
        # publish xls form
        self._publish_transportation_form_and_submit_instance()
        # create export db object
        export = generate_export(
            Export.CSV_ZIP_EXPORT, "zip", self.user.username,
            self.xform.id_string, group_delimiter='/',
            split_select_multiples=True)
        storage = get_storage_class()()
        self.assertTrue(storage.exists(export.filepath))
        path, ext = os.path.splitext(export.filename)
        self.assertEqual(ext, '.zip')


class TestExportBuilder(MainTestCase):
    data = [
        {
            'name': 'Abe',
            'age': 35,
            'tel/telLg==office': '020123456',
            'children':
            [
                {
                    'children/name': 'Mike',
                    'children/age': 5,
                    'children/fav_colors': 'red blue',
                    'children/iceLg==creams': 'vanilla chocolate',
                    'children/cartoons':
                    [
                        {
                            'children/cartoons/name': 'Tom & Jerry',
                            'children/cartoons/why': 'Tom is silly',
                        },
                        {
                            'children/cartoons/name': 'Flinstones',
                            'children/cartoons/why': u"I like bam bam\u0107"
                            # throw in a unicode character
                        }
                    ]
                },
                {
                    'children/name': 'John',
                    'children/age': 2,
                    'children/cartoons': []
                },
                {
                    'children/name': 'Imora',
                    'children/age': 3,
                    'children/cartoons':
                    [
                        {
                            'children/cartoons/name': 'Shrek',
                            'children/cartoons/why': 'He\'s so funny'
                        },
                        {
                            'children/cartoons/name': 'Dexter\'s Lab',
                            'children/cartoons/why': 'He thinks hes smart',
                            'children/cartoons/characters':
                            [
                                {
                                    'children/cartoons/characters/name': 'Dee Dee',
                                    'children/cartoons/characters/good_or_evil': 'good'
                                },
                                {
                                    'children/cartoons/characters/name': 'Dexter',
                                    'children/cartoons/characters/good_or_evil': 'evil'
                                },
                            ]
                        }
                    ]
                }
            ]
        },
        {
            # blank data just to be sure
            'children': []
        }
    ]
    long_survey_data = [
        {
            'name': 'Abe',
            'age': 35,
            'childrens_survey_with_a_very_lo':
            [
                {
                    'childrens_survey_with_a_very_lo/name': 'Mike',
                    'childrens_survey_with_a_very_lo/age': 5,
                    'childrens_survey_with_a_very_lo/fav_colors': 'red blue',
                    'childrens_survey_with_a_very_lo/cartoons':
                    [
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name': 'Tom & Jerry',
                            'childrens_survey_with_a_very_lo/cartoons/why': 'Tom is silly',
                        },
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name': 'Flinstones',
                            'childrens_survey_with_a_very_lo/cartoons/why': u"I like bam bam\u0107"
                            # throw in a unicode character
                        }
                    ]
                },
                {
                    'childrens_survey_with_a_very_lo/name': 'John',
                    'childrens_survey_with_a_very_lo/age': 2,
                    'childrens_survey_with_a_very_lo/cartoons': []
                },
                {
                    'childrens_survey_with_a_very_lo/name': 'Imora',
                    'childrens_survey_with_a_very_lo/age': 3,
                    'childrens_survey_with_a_very_lo/cartoons':
                    [
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name': 'Shrek',
                            'childrens_survey_with_a_very_lo/cartoons/why': 'He\'s so funny'
                        },
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name': 'Dexter\'s Lab',
                            'childrens_survey_with_a_very_lo/cartoons/why': 'He thinks hes smart',
                            'childrens_survey_with_a_very_lo/cartoons/characters':
                            [
                                {
                                    'childrens_survey_with_a_very_lo/cartoons/characters/name': 'Dee Dee',
                                    'children/cartoons/characters/good_or_evil': 'good'
                                },
                                {
                                    'childrens_survey_with_a_very_lo/cartoons/characters/name': 'Dexter',
                                    'children/cartoons/characters/good_or_evil': 'evil'
                                },
                            ]
                        }
                    ]
                }
            ]
        }
    ]
    data_utf8 = [
        {
            'name': 'Abe',
            'age': 35,
            'tel/telLg==office': '020123456',
            'childrenLg==info':
            [
                {
                    'childrenLg==info/nameLg==first': 'Mike',
                    'childrenLg==info/age': 5,
                    'childrenLg==info/fav_colors': u'red\u2019s blue\u2019s',
                    'childrenLg==info/ice_creams': 'vanilla chocolate',
                    'childrenLg==info/cartoons':
                    [
                        {
                            'childrenLg==info/cartoons/name': 'Tom & Jerry',
                            'childrenLg==info/cartoons/why': 'Tom is silly',
                        },
                        {
                            'childrenLg==info/cartoons/name': 'Flinstones',
                            'childrenLg==info/cartoons/why': u"I like bam bam\u0107"
                            # throw in a unicode character
                        }
                    ]
                }
            ]
        }
    ]

    def _create_childrens_survey(self):
        survey = create_survey_from_xls(
            os.path.join(
                os.path.abspath('./'), 'odk_logger', 'tests', 'fixtures',
                'childrens_survey.xls'))
        return survey

    def test_build_sections_from_survey(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        # test that we generate the proper sections
        expected_sections = [
            survey.name, 'children', 'children/cartoons',
            'children/cartoons/characters']
        self.assertEqual(
            expected_sections, [s['name'] for s in export_builder.sections])
        # main section should have split geolocations
        expected_element_names = [
            'name', 'age', 'geo/geolocation', 'geo/_geolocation_longitude',
            'geo/_geolocation_latitude', 'geo/_geolocation_altitude',
            'geo/_geolocation_precision', 'tel/tel.office', 'tel/tel.mobile',
            'meta/instanceID']
        section = export_builder.section_by_name(survey.name)
        element_names = [element['xpath'] for element in section['elements']]
        # fav_colors should have its choices split
        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

        expected_element_names = [
            'children/name', 'children/age', 'children/fav_colors',
            'children/fav_colors/red', 'children/fav_colors/blue',
            'children/fav_colors/pink', 'children/ice.creams',
            'children/ice.creams/vanilla', 'children/ice.creams/strawberry',
            'children/ice.creams/chocolate']
        section = export_builder.section_by_name('children')
        element_names = [element['xpath'] for element in section['elements']]
        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

        expected_element_names = [
            'children/cartoons/name', 'children/cartoons/why']
        section = export_builder.section_by_name('children/cartoons')
        element_names = [element['xpath'] for element in section['elements']]

        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

        expected_element_names = [
            'children/cartoons/characters/name',
            'children/cartoons/characters/good_or_evil']
        section = export_builder.section_by_name('children/cartoons/characters')
        element_names = [element['xpath'] for element in section['elements']]
        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

    def test_zipped_csv_export_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_csv(temp_zip_file.name, self.data)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()

        # generate data to compare with
        index = 1
        indices = {}
        survey_name = survey.name
        outputs = []
        for d in self.data:
            outputs.append(
                dict_to_joined_export(d, index, indices, survey_name))
            index += 1

        # check that each file exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "{0}.csv".format(survey.name))))
        with open(
                os.path.join(
                    temp_dir, "{0}.csv".format(survey.name))) as csv_file:
            reader = csv.reader(csv_file)
            rows = [r for r in reader]

            # open comparison file
            with open(
                os.path.join(
                    os.path.abspath('./'), 'odk_logger', 'tests', 'fixtures',
                    'csvs', 'childrens_survey.csv')) as fixture_csv:
                fixture_reader = csv.reader(fixture_csv)
                expected_rows = [r for r in fixture_reader]
                self.assertEqual(rows, expected_rows)

        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "children.csv")))
        with open(os.path.join(temp_dir, "children.csv")) as csv_file:
            reader = csv.reader(csv_file)
            rows = [r for r in reader]

            # open comparison file
            with open(
                os.path.join(
                    os.path.abspath('./'), 'odk_logger', 'tests', 'fixtures',
                    'csvs', 'children.csv')) as fixture_csv:
                fixture_reader = csv.reader(fixture_csv)
                expected_rows = [r for r in fixture_reader]
                self.assertEqual(rows, expected_rows)

        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "children_cartoons.csv")))
        with open(os.path.join(temp_dir, "children_cartoons.csv")) as csv_file:
            reader = csv.reader(csv_file)
            rows = [r for r in reader]

            # open comparison file
            with open(
                os.path.join(
                    os.path.abspath('./'), 'odk_logger', 'tests', 'fixtures',
                    'csvs', 'children_cartoons.csv')) as fixture_csv:
                fixture_reader = csv.reader(fixture_csv)
                expected_rows = [r for r in fixture_reader]
                self.assertEqual(rows, expected_rows)

        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "children_cartoons_characters.csv")))
        with open(os.path.join(
                temp_dir, "children_cartoons_characters.csv")) as csv_file:
            reader = csv.reader(csv_file)
            rows = [r for r in reader]

            # open comparison file
            with open(
                os.path.join(
                    os.path.abspath('./'), 'odk_logger', 'tests', 'fixtures',
                    'csvs', 'children_cartoons_characters.csv')) as fixture_csv:
                fixture_reader = csv.reader(fixture_csv)
                expected_rows = [r for r in fixture_reader]
                self.assertEqual(rows, expected_rows)

        shutil.rmtree(temp_dir)

    def test_decode_mongo_encoded_section_names(self):
        data = {
            'main_section': [1, 2, 3, 4],
            'sectionLg==1/info': [1, 2, 3, 4],
            'sectionLg==2/info': [1, 2, 3, 4],
        }
        result = ExportBuilder.decode_mongo_encoded_section_names(data)
        expected_result = {
            'main_section': [1, 2, 3, 4],
            'section.1/info': [1, 2, 3, 4],
            'section.2/info': [1, 2, 3, 4],
        }
        self.assertEqual(result, expected_result)

    def test_zipped_csv_export_works_with_unicode(self):
        """
        cvs writer doesnt handle unicode we we have to encode to ascii
        """
        survey = create_survey_from_xls(
            os.path.join(
                os.path.abspath('./'), 'odk_logger', 'tests', 'fixtures',
                'childrens_survey_unicode.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_csv(temp_zip_file.name, self.data_utf8)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "children.info.csv")))
        # check file's contents
        with open(os.path.join(temp_dir, "children.info.csv")) as csv_file:
            reader = csv.reader(csv_file)
            expected_headers = ['children.info/name.first',
                                'children.info/age',
                                'children.info/fav_colors',
                                u'children.info/fav_colors/red\u2019s',
                                u'children.info/fav_colors/blue\u2019s',
                                u'children.info/fav_colors/pink\u2019s',
                                'children.info/ice_creams',
                                'children.info/ice_creams/vanilla',
                                'children.info/ice_creams/strawberry',
                                'children.info/ice_creams/chocolate', '_id',
                                '_uuid', '_submission_time', '_index',
                                '_parent_table_name', '_parent_index']
            rows = [row for row in reader]
            actual_headers = [h.decode('utf-8') for h in rows[0]]
            self.assertEqual(sorted(actual_headers), sorted(expected_headers))
            data = dict(zip(rows[0], rows[1]))
            self.assertEqual(
                data[u'children.info/fav_colors/red\u2019s'.encode('utf-8')],
                'True')
            self.assertEqual(
                data[u'children.info/fav_colors/blue\u2019s'.encode('utf-8')],
                'True')
            self.assertEqual(
                data[u'children.info/fav_colors/pink\u2019s'.encode('utf-8')],
                'False')
            # check that red and blue are set to true
        shutil.rmtree(temp_dir)

    def test_xls_export_works_with_unicode(self):
        survey = create_survey_from_xls(
            os.path.join(
                os.path.abspath('./'), 'odk_logger', 'tests', 'fixtures',
                'childrens_survey_unicode.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix='.xlsx')
        export_builder.to_xls_export(temp_xls_file.name, self.data_utf8)
        temp_xls_file.seek(0)
        # check that values for red\u2019s and blue\u2019s are set to true
        wb = load_workbook(temp_xls_file.name)
        children_sheet = wb.get_sheet_by_name("children.info")
        data = dict([(r[0].value, r[1].value) for r in children_sheet.columns])
        self.assertTrue(data[u'children.info/fav_colors/red\u2019s'])
        self.assertTrue(data[u'children.info/fav_colors/blue\u2019s'])
        self.assertFalse(data[u'children.info/fav_colors/pink\u2019s'])
        temp_xls_file.close()

    def test_generation_of_multi_selects_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_select_multiples =\
            {
                'children':
                {
                    'children/fav_colors':
                    [
                        'children/fav_colors/red', 'children/fav_colors/blue',
                        'children/fav_colors/pink'
                    ],
                    'children/ice.creams':
                    [
                        'children/ice.creams/vanilla',
                        'children/ice.creams/strawberry',
                        'children/ice.creams/chocolate'
                    ]
                }
            }
        select_multiples = export_builder.select_multiples
        self.assertTrue('children' in select_multiples)
        self.assertTrue('children/fav_colors' in select_multiples['children'])
        self.assertTrue('children/ice.creams' in select_multiples['children'])
        self.assertEqual(
            sorted(select_multiples['children']['children/fav_colors']),
            sorted(
                expected_select_multiples['children']['children/fav_colors']))
        self.assertEqual(
            sorted(select_multiples['children']['children/ice.creams']),
            sorted(
                expected_select_multiples['children']['children/ice.creams']))

    def test_split_select_multiples_works(self):
        select_multiples =\
            {
                'children/fav_colors': [
                    'children/fav_colors/red', 'children/fav_colors/blue',
                    'children/fav_colors/pink']
            }
        row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': 'red blue'
            }
        new_row = ExportBuilder.split_select_multiples(
            row, select_multiples)
        expected_row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': 'red blue',
                'children/fav_colors/red': True,
                'children/fav_colors/blue': True,
                'children/fav_colors/pink': False
            }
        self.assertEqual(new_row, expected_row)

    def test_split_select_multiples_works_when_data_is_blank(self):
        select_multiples =\
            {
                'children/fav_colors': [
                    'children/fav_colors/red', 'children/fav_colors/blue',
                    'children/fav_colors/pink']
            }
        row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': ''
            }
        new_row = ExportBuilder.split_select_multiples(
            row, select_multiples)
        expected_row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': '',
                'children/fav_colors/red': False,
                'children/fav_colors/blue': False,
                'children/fav_colors/pink': False
            }
        self.assertEqual(new_row, expected_row)

    def test_generation_of_gps_fields_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_gps_fields =\
            {
                'childrens_survey':
                {
                    'geo/geolocation':
                    [
                        'geo/_geolocation_latitude', 'geo/_geolocation_longitude',
                        'geo/_geolocation_altitude', 'geo/_geolocation_precision'
                    ]
                }
            }
        gps_fields = export_builder.gps_fields
        self.assertTrue(gps_fields.has_key('childrens_survey'))
        self.assertEqual(
            sorted(gps_fields['childrens_survey']),
            sorted(expected_gps_fields['childrens_survey']))

    def test_split_gps_components_works(self):
        gps_fields =\
            {
                'geo/geolocation':
                [
                    'geo/_geolocation_latitude', 'geo/_geolocation_longitude',
                    'geo/_geolocation_altitude', 'geo/_geolocation_precision'
                ]
            }
        row = \
            {
                'geo/geolocation': '1.0 36.1 2000 20',
            }
        new_row = ExportBuilder.split_gps_components(
            row, gps_fields)
        expected_row = \
            {
                'geo/geolocation': '1.0 36.1 2000 20',
                'geo/_geolocation_latitude': '1.0',
                'geo/_geolocation_longitude': '36.1',
                'geo/_geolocation_altitude': '2000',
                'geo/_geolocation_precision': '20'
            }
        self.assertEqual(new_row, expected_row)

    def test_split_gps_components_works_when_gps_data_is_blank(self):
        gps_fields =\
            {
                'geo/geolocation':
                [
                    'geo/_geolocation_latitude', 'geo/_geolocation_longitude',
                    'geo/_geolocation_altitude', 'geo/_geolocation_precision'
                ]
            }
        row = \
            {
                'geo/geolocation': '',
            }
        new_row = ExportBuilder.split_gps_components(
            row, gps_fields)
        expected_row = \
            {
                'geo/geolocation': '',
            }
        self.assertEqual(new_row, expected_row)

    def test_generation_of_mongo_encoded_fields_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_encoded_fields =\
            {
                'childrens_survey':
                {
                    'tel/tel.office': 'tel/{0}'.format(
                        _encode_for_mongo('tel.office')),
                    'tel/tel.mobile': 'tel/{0}'.format(
                        _encode_for_mongo('tel.mobile')),
                }
            }
        encoded_fields = export_builder.encoded_fields
        self.assertTrue('childrens_survey' in encoded_fields)
        self.assertEqual(
            encoded_fields['childrens_survey'],
            expected_encoded_fields['childrens_survey'])

    def test_decode_fields_names_encoded_for_mongo(self):
        encoded_fields = \
            {
                'tel/tel.office': 'tel/{0}'.format(
                    _encode_for_mongo('tel.office'))
            }
        row = \
            {
                'name': 'Abe',
                'age': 35,
                'tel/{0}'.format(_encode_for_mongo('tel.office')): '123-456-789'
            }
        new_row = ExportBuilder.decode_mongo_encoded_fields(row, encoded_fields)
        expected_row = \
            {
                'name': 'Abe',
                'age': 35,
                'tel/tel.office': '123-456-789'
            }
        self.assertEqual(new_row, expected_row)

    def test_generate_field_title(self):
        field_name = ExportBuilder.format_field_title("child/age", ".")
        expected_field_name = "child.age"
        self.assertEqual(field_name, expected_field_name)

    def test_delimiter_replacement_works_existing_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_sections =\
            [
                {
                    'name': 'children',
                    'elements': [
                        {
                            'title': 'children.name',
                            'xpath': 'children/name'
                        }
                    ]
                }
            ]
        children_section = export_builder.section_by_name('children')
        self.assertEqual(
            children_section['elements'][0]['title'],
            expected_sections[0]['elements'][0]['title'])

    def test_delimiter_replacement_works_generated_multi_select_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_section =\
            {
                'name': 'children',
                'elements': [
                    {
                        'title': 'children.fav_colors.red',
                        'xpath': 'children/fav_colors/red'
                    }
                ]
            }
        childrens_section = export_builder.section_by_name('children')
        match = filter(lambda x: expected_section['elements'][0]['xpath']
                       == x['xpath'], childrens_section['elements'])[0]
        self.assertEqual(
            expected_section['elements'][0]['title'], match['title'])

    def test_delimiter_replacement_works_for_generated_gps_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_section = \
            {
                'name': 'childrens_survey',
                'elements': [
                    {
                        'title': 'geo._geolocation_latitude',
                        'xpath': 'geo/_geolocation_latitude'
                    }
                ]
            }
        main_section = export_builder.section_by_name('childrens_survey')
        match = filter(
            lambda x: (expected_section['elements'][0]['xpath']
                       == x['xpath']), main_section['elements'])[0]
        self.assertEqual(
            expected_section['elements'][0]['title'], match['title'])

    def test_to_xls_export_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xls')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.data)
        xls_file.seek(0)
        wb = load_workbook(filename)
        # check that we have childrens_survey, children, children_cartoons
        # and children_cartoons_characters sheets
        expected_sheet_names = ['childrens_survey', 'children',
                                'children_cartoons',
                                'children_cartoons_characters']
        self.assertEqual(wb.get_sheet_names(), expected_sheet_names)

        # check header columns
        main_sheet = wb.get_sheet_by_name('childrens_survey')
        expected_column_headers = [
            u'name', u'age', u'geo/geolocation', u'geo/_geolocation_latitude',
            u'geo/_geolocation_longitude', u'geo/_geolocation_altitude',
            u'geo/_geolocation_precision', u'tel/tel.office',
            u'tel/tel.mobile', u'_id', u'meta/instanceID', u'_uuid',
            u'_submission_time', u'_index', u'_parent_index',
            u'_parent_table_name']
        column_headers = [c[0].value for c in main_sheet.columns]
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        childrens_sheet = wb.get_sheet_by_name('children')
        expected_column_headers = [
            u'children/name', u'children/age', u'children/fav_colors',
            u'children/fav_colors/red', u'children/fav_colors/blue',
            u'children/fav_colors/pink', u'children/ice.creams',
            u'children/ice.creams/vanilla', u'children/ice.creams/strawberry',
            u'children/ice.creams/chocolate', u'_id', u'_uuid',
            u'_submission_time', u'_index', u'_parent_index',
            u'_parent_table_name']
        column_headers = [c[0].value for c in childrens_sheet.columns]
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        cartoons_sheet = wb.get_sheet_by_name('children_cartoons')
        expected_column_headers = [
            u'children/cartoons/name', u'children/cartoons/why', u'_id',
            u'_uuid', u'_submission_time', u'_index', u'_parent_index',
            u'_parent_table_name']
        column_headers = [c[0].value for c in cartoons_sheet.columns]
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        characters_sheet = wb.get_sheet_by_name('children_cartoons_characters')
        expected_column_headers = [
            u'children/cartoons/characters/name',
            u'children/cartoons/characters/good_or_evil', u'_id', u'_uuid',
            u'_submission_time', u'_index', u'_parent_index',
            u'_parent_table_name']
        column_headers = [c[0].value for c in characters_sheet.columns]
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        xls_file.close()

    def test_to_xls_export_respects_custom_field_delimiter(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = ExportBuilder.GROUP_DELIMITER_DOT
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xls')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.data)
        xls_file.seek(0)
        wb = load_workbook(filename)

        # check header columns
        main_sheet = wb.get_sheet_by_name('childrens_survey')
        expected_column_headers = [
            u'name', u'age', u'geo.geolocation', u'geo._geolocation_latitude',
            u'geo._geolocation_longitude', u'geo._geolocation_altitude',
            u'geo._geolocation_precision', u'tel.tel.office',
            u'tel.tel.mobile', u'_id', u'meta.instanceID', u'_uuid',
            u'_submission_time', u'_index', u'_parent_index',
            u'_parent_table_name']
        column_headers = [c[0].value for c in main_sheet.columns]
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))
        xls_file.close()

    def test_get_valid_sheet_name_catches_duplicates(self):
        work_sheets = {'childrens_survey': "Worksheet"}
        desired_sheet_name = "childrens_survey"
        expected_sheet_name = "childrens_survey1"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, work_sheets)
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    def test_get_valid_sheet_name_catches_long_names(self):
        desired_sheet_name = "childrens_survey_with_a_very_long_name"
        expected_sheet_name = "childrens_survey_with_a_very_lo"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, [])
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    def test_get_valid_sheet_name_catches_long_duplicate_names(self):
        work_sheet_titles = ['childrens_survey_with_a_very_lo']
        desired_sheet_name = "childrens_survey_with_a_very_long_name"
        expected_sheet_name = "childrens_survey_with_a_very_l1"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, work_sheet_titles)
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    def test_to_xls_export_generates_valid_sheet_names(self):
        survey = create_survey_from_xls(
            os.path.join(
                os.path.abspath('./'), 'odk_logger', 'tests', 'fixtures',
                'childrens_survey_with_a_very_long_name.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xls')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.data)
        xls_file.seek(0)
        wb = load_workbook(filename)
        # check that we have childrens_survey, children, children_cartoons
        # and children_cartoons_characters sheets
        expected_sheet_names = ['childrens_survey_with_a_very_lo',
                                'childrens_survey_with_a_very_l1',
                                'childrens_survey_with_a_very_l2',
                                'childrens_survey_with_a_very_l3']
        self.assertEqual(wb.get_sheet_names(), expected_sheet_names)
        xls_file.close()

    def test_child_record_parent_table_is_updated_when_sheet_is_renamed(self):
        survey = create_survey_from_xls(
            os.path.join(
                os.path.abspath('./'), 'odk_logger', 'tests', 'fixtures',
                'childrens_survey_with_a_very_long_name.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xlsx')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.long_survey_data)
        xls_file.seek(0)
        wb = load_workbook(filename)

        # get the children's sheet
        ws1 = wb.get_sheet_by_name('childrens_survey_with_a_very_l1')

        # parent_table is in cell K2
        parent_table_name = ws1.cell('K2').value
        expected_parent_table_name = 'childrens_survey_with_a_very_lo'
        self.assertEqual(parent_table_name, expected_parent_table_name)

        # get cartoons sheet
        ws2 = wb.get_sheet_by_name('childrens_survey_with_a_very_l2')
        parent_table_name = ws2.cell('G2').value
        expected_parent_table_name = 'childrens_survey_with_a_very_l1'
        self.assertEqual(parent_table_name, expected_parent_table_name)
        xls_file.close()

    def test_type_conversion(self):
        submission_1 = {
            "_id": 579827,
            "geolocation": "-1.2625482 36.7924794 0.0 21.0",
            "_bamboo_dataset_id": "",
            "meta/instanceID": "uuid:2a8129f5-3091-44e1-a579-bed2b07a12cf",
            "name": "Smith",
            "formhub/uuid": "633ec390e024411ba5ce634db7807e62",
            "_submission_time": "2013-07-03T08:25:30",
            "age": "107",
            "_uuid": "2a8129f5-3091-44e1-a579-bed2b07a12cf",
            "when": "2013-07-03",
#            "_deleted_at": None,
            "amount": "250.0",
            "_geolocation": [
                "-1.2625482",
                "36.7924794"
            ],
            "_xform_id_string": "test_data_types",
            "_userform_id": "larryweya_test_data_types",
            "_status": "submitted_via_web",
            "precisely": "2013-07-03T15:24:00.000+03",
            "really": "15:24:00.000+03"
        }

        submission_2 = {
            "_id": 579828,
            "_submission_time": "2013-07-03T08:26:10",
            "_uuid": "5b4752eb-e13c-483e-87cb-e67ca6bb61e5",
            "_bamboo_dataset_id": "",
#            "_deleted_at": None,
            "_xform_id_string": "test_data_types",
            "_userform_id": "larryweya_test_data_types",
            "_status": "submitted_via_web",
            "meta/instanceID": "uuid:5b4752eb-e13c-483e-87cb-e67ca6bb61e5",
            "formhub/uuid": "633ec390e024411ba5ce634db7807e62",
            "amount": "",
        }

        survey = create_survey_from_xls(
            os.path.join(
                os.path.abspath('./'), 'odk_viewer', 'tests', 'fixtures',
                'test_data_types/test_data_types.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        # format submission 1 for export
        survey_name = survey.name
        indices = {survey_name: 0}
        data = dict_to_joined_export(submission_1, 1, indices, survey_name)
        new_row = export_builder.pre_process_row(data[survey_name],
                                                 export_builder.sections[0])
        self.assertIsInstance(new_row['age'], int)
        self.assertIsInstance(new_row['when'], datetime.date)
        #self.assertIsInstance(new_row['precisely'], datetime.datetime)
        self.assertIsInstance(new_row['amount'], float)
        #self.assertIsInstance(new_row['_submission_time'], datetime.datetime)
        #self.assertIsInstance(new_row['really'], datetime.time)

        # check missing values dont break and empty values return blank strings
        indices = {survey_name: 0}
        data = dict_to_joined_export(submission_2, 1, indices, survey_name)
        new_row = export_builder.pre_process_row(data[survey_name],
                                                 export_builder.sections[0])
        self.assertIsInstance(new_row['amount'], basestring)
        self.assertEqual(new_row['amount'], '')

    def test_xls_convert_dates_before_1900(self):
        survey = create_survey_from_xls(
            os.path.join(
                os.path.abspath('./'), 'odk_viewer', 'tests', 'fixtures',
                'test_data_types/test_data_types.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        data = [
            {
                'name': 'Abe',
                'when': '1899-07-03',
            }
        ]
        # create export file
        temp_xls_file = NamedTemporaryFile(suffix='.xlsx')
        export_builder.to_xls_export(temp_xls_file.name, data)
        temp_xls_file.close()
        # this should error if there is a problem, not sure what to assert

    def test_convert_types(self):
        val = '1'
        expected_val = 1
        converted_val = ExportBuilder.convert_type(val, 'int')
        self.assertIsInstance(converted_val, int)
        self.assertEqual(converted_val, expected_val)

        val = '1.2'
        expected_val = 1.2
        converted_val = ExportBuilder.convert_type(val, 'decimal')
        self.assertIsInstance(converted_val, float)
        self.assertEqual(converted_val, expected_val)

        val = '2012-06-23'
        expected_val = datetime.date(2012, 6, 23)
        converted_val = ExportBuilder.convert_type(val, 'date')
        self.assertIsInstance(converted_val, datetime.date)
        self.assertEqual(converted_val, expected_val)
